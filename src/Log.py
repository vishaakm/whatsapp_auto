import openpyxl
import datetime

class Log:
    def __init__(self):
        self.rowCount = 2
        # Create workbook
        self.wb = openpyxl.Workbook()
        # Select active worksheet
        self.ws = self.wb.active
        # Set values to first row
        self.ws['A1'] = 'Date'
        self.ws['B1'] = 'Name'
        self.ws['C1'] = 'Contact'
        self.ws['D1'] = 'Process'

    def writeLog(self,name,contact,process):
        self.ws['A{0}'.format(self.rowCount)] = datetime.datetime.now()
        self.ws['B{0}'.format(self.rowCount)] = name
        self.ws['C{0}'.format(self.rowCount)] = contact
        self.ws['D{0}'.format(self.rowCount)] = process
        self.wb.save('Log.xlsx')
        self.rowCount += 1






